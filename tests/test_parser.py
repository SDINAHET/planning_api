from app.parser import clean, build_mission_and_lieu, get_schedule, extract_general
from openpyxl import Workbook
from app.parser import load_people, extract_all
from io import BytesIO
from app.parser import open_workbook

def test_clean():
    assert clean(None) == ""
    assert clean(" hello\nworld ") == "hello world"


def test_build_mission_surveillance_salle():
    mission, lieu = build_mission_and_lieu("Surveillance salle", "Amphi A")
    assert mission == "Surveillance salle"
    assert lieu == "Amphi A"


def test_build_mission_keynote():
    mission, lieu = build_mission_and_lieu("Keynote ouverture", "")
    assert mission == "Keynote"
    assert lieu == "Amphi principal"


def test_build_mission_accueil():
    mission, lieu = build_mission_and_lieu("Accueil", "")
    assert mission == "Accueil"
    assert lieu == "Accueil"


def test_get_schedule_with_mock(monkeypatch):
    def fake_extract_all(path):
        return {
            "Stéphane DINAHET": [{"mission": "Accueil"}],
            "Autre Personne": [{"mission": "Salle"}],
        }

    monkeypatch.setattr("app.parser.extract_all", fake_extract_all)

    result = get_schedule("fake.xlsx", "Stéphane")
    assert len(result) == 1
    assert result[0]["mission"] == "Accueil"


def test_get_schedule_not_found(monkeypatch):
    monkeypatch.setattr("app.parser.extract_all", lambda path: {"Jean Test": []})

    assert get_schedule("fake.xlsx", "Inconnu") == []


def test_extract_general_with_mock(monkeypatch):
    def fake_extract_all(path):
        return {
            "Alice": [
                {
                    "jour": "Mercredi",
                    "date": "2026-06-24",
                    "horaire": "10h-12h",
                    "mission": "Accueil",
                    "lieu": "Accueil",
                    "benevoles": ["Alice"],
                    "nb_benevoles": 1,
                }
            ],
            "Bob": [
                {
                    "jour": "Mercredi",
                    "date": "2026-06-24",
                    "horaire": "10h-12h",
                    "mission": "Accueil",
                    "lieu": "Accueil",
                    "benevoles": ["Bob"],
                    "nb_benevoles": 1,
                }
            ],
        }

    monkeypatch.setattr("app.parser.extract_all", fake_extract_all)

    result = extract_general("fake.xlsx")

    assert len(result) == 1
    assert result[0]["nb_benevoles"] == 2
    assert "Alice" in result[0]["benevoles"]
    assert "Bob" in result[0]["benevoles"]



def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"

    ws.cell(row=3, column=9).value = "Alice"
    ws.cell(row=3, column=10).value = "Bob"

    ws.cell(row=4, column=1).value = "Mercredi"
    ws.cell(row=4, column=3).value = "10h-12h"
    ws.cell(row=4, column=4).value = "Accueil"
    ws.cell(row=4, column=9).value = "x"
    ws.cell(row=4, column=10).value = "x"

    return wb


def test_load_people(monkeypatch):
    wb = create_workbook()

    monkeypatch.setattr(
        "app.parser.open_workbook",
        lambda _: wb
    )

    people = load_people("fake.xlsx")

    assert people == ["Alice", "Bob"]


def test_extract_all(monkeypatch):
    wb = create_workbook()

    monkeypatch.setattr(
        "app.parser.open_workbook",
        lambda _: wb
    )

    schedules = extract_all("fake.xlsx")

    assert "Alice" in schedules
    assert "Bob" in schedules

    assert len(schedules["Alice"]) == 1

    slot = schedules["Alice"][0]

    assert slot["jour"] == "Mercredi"
    assert slot["horaire"] == "10h-12h"
    assert slot["mission"] == "Accueil"
    assert slot["nb_benevoles"] == 2


def test_open_workbook_local(monkeypatch):
    wb = create_workbook()

    monkeypatch.setattr(
        "app.parser.load_workbook",
        lambda *a, **k: wb
    )

    result = open_workbook("fake.xlsx")

    assert result == wb


def test_open_workbook_http(monkeypatch):
    wb = create_workbook()

    class FakeResponse:
        content = b"123"

        def raise_for_status(self):
            pass

    monkeypatch.setattr(
        "app.parser.requests.get",
        lambda *a, **k: FakeResponse()
    )

    monkeypatch.setattr(
        "app.parser.load_workbook",
        lambda *a, **k: wb
    )

    result = open_workbook("https://test")

    assert result == wb
