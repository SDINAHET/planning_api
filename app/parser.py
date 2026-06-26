from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Any
from openpyxl import load_workbook
from io import BytesIO
import requests

DEFAULT_DAY_DATES = {
    "Mardi": "2026-06-23",
    "Mercredi": "2026-06-24",
    "Jeudi": "2026-06-25",
    "Vendredi": "2026-06-26",
}

HEADER_ROW = 3
FIRST_PERSON_COL = 9
DAY_COL = 1
TIME_COL = 3
TASK_COL = 4
SUBTASK_COL = 5


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def open_workbook(source: str | Path):
    source = str(source)

    if source.startswith("http"):
        response = requests.get(source, timeout=15)
        response.raise_for_status()
        return load_workbook(BytesIO(response.content), data_only=True)

    return load_workbook(source, data_only=True)


def load_people(path: str | Path) -> List[str]:
    wb = open_workbook(path)
    ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active

    people = []

    for col in range(FIRST_PERSON_COL, ws.max_column + 1):
        name = clean(ws.cell(HEADER_ROW, col).value)

        if name and name not in ("41",) and not name.isdigit():
            people.append(name)

    return people


def extract_all(path: str | Path) -> Dict[str, List[dict]]:
    wb = open_workbook(path)
    ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active

    people_cols = {}

    for col in range(FIRST_PERSON_COL, ws.max_column + 1):
        name = clean(ws.cell(HEADER_ROW, col).value)

        if name and not name.isdigit():
            people_cols[col] = name

    schedules: Dict[str, List[dict]] = {
        name: [] for name in people_cols.values()
    }

    current_day = ""
    current_time = ""
    current_main_task = ""

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        day = clean(ws.cell(row, DAY_COL).value).replace('"', "").strip()
        time = clean(ws.cell(row, TIME_COL).value)
        main_task = clean(ws.cell(row, TASK_COL).value)
        sub_task = clean(ws.cell(row, SUBTASK_COL).value)

        if day:
            current_day = day

        if time:
            current_time = time

        if main_task:
            current_main_task = main_task

        if not current_day or not current_time:
            continue

        if (
            "Nombre de créneaux" in current_main_task
            or "Total créneaux" in current_main_task
        ):
            continue

        mission = sub_task or current_main_task

        lieu = (
            sub_task
            if sub_task.startswith("Amphi")
            else ("Accueil" if "Accueil" in mission else "")
        )

        if current_main_task.lower().startswith("surveillance salle") and sub_task:
            mission = "Surveillance salle"
            lieu = sub_task

        elif "Keynote" in current_main_task:
            mission = "Keynote"
            lieu = sub_task or "Amphi principal"

        elif sub_task and current_main_task and not sub_task.startswith("Amphi"):
            mission = f"{current_main_task} - {sub_task}"

        for col, person in people_cols.items():
            marker = clean(ws.cell(row, col).value).lower()

            if marker == "x":
                schedules[person].append(
                    {
                        "jour": current_day,
                        "date": DEFAULT_DAY_DATES.get(current_day, ""),
                        "horaire": current_time.replace(" ", ""),
                        "mission": mission,
                        "lieu": lieu,
                    }
                )

    return schedules


def get_schedule(path: str | Path, person: str) -> List[dict]:
    all_schedules = extract_all(path)
    target = person.lower().strip()

    for name, planning in all_schedules.items():
        if target in name.lower() or name.lower() in target:
            return planning

    return []
